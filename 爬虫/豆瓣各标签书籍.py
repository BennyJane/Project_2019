#有些问题待处理
import sys
import time
import urllib
#import urllib2
import requests
import numpy as np
from bs4 import BeautifulSoup
from openpyxl import Workbook
from multiprocessing import Pool
import time

#reload(sys)
#sys.setdefaultencoding('utf8')

# Some User Agents
hds = [{'User-Agent': 'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6'}, \
       {
           'User-Agent': 'Mozilla/5.0 (Windows NT 6.2) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.12 Safari/535.11'}, \
       {'User-Agent': 'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Trident/6.0)'}]


def book_spider(book_tag):
    page_num = 0
    book_list = []
    try_times = 0

    while (1):
        # url='http://www.douban.com/tag/%E5%B0%8F%E8%AF%B4/book?start=0' # For Test
        url = 'http://www.douban.com/tag/{}/book?start={}'.format(book_tag,str(page_num *15))
        time.sleep(np.random.rand() * 5)

        try:
            req = requests.get(url, headers=hds[page_num % len(hds)])
            source = req.text
        except Exception as  e:
            print(e)
            continue

        #soup = BeautifulSoup(source, 'html.parser', verify=False)
        soup = BeautifulSoup(source,'html.parser')
        list_soup = soup.find('div', {'class': 'mod book-list'})

        try_times += 1
        if list_soup == None and try_times < 200:
            continue
        elif list_soup == None or len(list_soup) <= 1:
            break  # Break when no informatoin got after 200 times requesting

        for book_info in list_soup.findAll('dd'):
            title = book_info.find('a', {'class': 'title'}).string.strip()
            desc = book_info.find('div', {'class': 'desc'}).string.strip()
            desc_list = desc.split('/')
            book_url = book_info.find('a', {'class': 'title'}).get('href')

            try:
                author_info = '作者/译者： ' + '/'.join(desc_list[0:-3])
            except:
                author_info = '作者/译者： 暂无'
            try:
                pub_info = '出版信息： ' + '/'.join(desc_list[-3:])
            except:
                pub_info = '出版信息： 暂无'
            try:
                rating = book_info.find('span', {'class': 'rating_nums'}).string.strip()
            except:
                rating = '0.0'
            try:
                # people_num = book_info.findAll('span')[2].string.strip()
                people_num = get_people_num(book_url)
                people_num = people_num.strip('人评价')
            except:
                people_num = '0'

            book_list.append([title, rating, people_num, author_info, pub_info])
            try_times = 0  # set 0 when got valid information
        page_num += 1
        print('Downloading Information From Page %d' % page_num)
        sys.stdout.flush()

    return book_list


def get_people_num(url):
    # url='http://book.douban.com/subject/6082808/?from=tag_all' # For Test
    global source
    try:
        req = requests.get(url, headers=hds[np.random.randint(0, len(hds))])
        source=req.text
    except Exception as e:
        print(e)

    #soup = BeautifulSoup(source ,"html.parser",verify=False)
    soup = BeautifulSoup(source, 'html.parser')
    people_num = soup.find('div', {'class': 'rating_sum'}).findAll('span')[1].string.strip()
    return people_num

def do_spider(book_tag_lists):
    book_lists = []
    for book_tag in book_tag_lists:
        book_list = book_spider(book_tag)
        book_list = sorted(book_list, key=lambda x: x[1], reverse=True)
        book_lists.append(book_list)
    return book_lists


def print_book_lists_excel(book_lists, book_tag_lists):
    #wb = Workbook(optimized_write=True)
    wb = Workbook(write_only=True)
    ws = []
    for i in range(len(book_tag_lists)):
        sheet=wb.create_sheet(title=book_tag_lists[i])  # utf8->unicode
        sheet.append(['序号', '书名', '评分', '评价人数', '作者', '出版社'])
        count = 1
        for bl in book_lists[i]:
            ws[i].append([count, bl[0], float(bl[1]), int(bl[2]), bl[3], bl[4]])
            count += 1
    save_path = 'E:/Python学习/Python Projects/爬虫项目331/book_list-{}'.format("-".join([item for item in book_tag_lists]))
    save_path += '.xlsx'
    #检查文件是否已存在
    if os.path.exists(save_path):
        os.remove(save_path)
    wb.save(save_path)

def Simple_tagbooks(book_tag_lists):
    start_time = time.time()
    book_lists = do_spider(book_tag_lists)
    print_book_lists_excel(book_lists, book_tag_lists)
    print("总计时长：",str(time.time()-start_time))



if __name__ == '__main__':
    # book_tag_lists = ['心理','判断与决策','算法','数据结构','经济','历史']
    # book_tag_lists = ['传记','哲学','编程','创业','理财','社会学','佛教']
    # book_tag_lists = ['思想','科技','科学','web','股票','爱情','两性']
    # book_tag_lists = ['计算机','机器学习','linux','android','数据库','互联网']
    # book_tag_lists = ['数学']
    # book_tag_lists = ['摄影','设计','音乐','旅行','教育','成长','情感','育儿','健康','养生']
    # book_tag_lists = ['商业','理财','管理']
    # book_tag_lists = ['名著']
    # book_tag_lists = ['科普','经典','生活','心灵','文学']
    # book_tag_lists = ['科幻','思维','金融']
    #book_tag_lists = ['个人管理', '时间管理', '投资', '文化', '宗教']
    book_tag_lists = ['时间管理']
    p=Pool(processes=4)
    p.apply_async(Simple_tagbooks,args=(book_tag_lists,))
    p.close()
    p.join()

'''
Error in atexit._run_exitfuncs:
Traceback (most recent call last):
  File "C:\Users\Administrator\Anaconda3\lib\site-packages\openpyxl\worksheet\write_only.py", line 33, in _openpyxl_shutdown
    os.remove(path)
PermissionError: [WinError 32] 另一个程序正在使用此文件，进程无法访问。: 'C:\\Users\\ADMINI~1\\AppData\\Local\\Temp\\openpyxl.k3rcw6lj'
'''
